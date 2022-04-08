import openpyxl
import csv
class userOperations():
    def userFunctions(self):
        print("Select operation you want to perform :\n1.Login \n2.Register \3.Exit")
        choise=int(input())
        match choise:
            case 1:
                while(True):
                    print("***^ Welcome to Login Page ^***")
                    UserName = input("Enter your user name :")
                    Password = str(input("Enter the Password :"))
                    a = userOperations()
                    if a.Login(UserName, Password):
                        print("Logged in Successfully")
                        a.LoginFunctionalites()
                        break
                    else:
                        print("Login failed")
            case 2:
                a=userOperations()
                a.register()
            case 3:
                pass

    def register(self):
        path = "data.xlsx"
        lst = []
        print("****Create new Account***** ")
        lst.append(input("Enter Name : "))
        lst.append(input("Enter Email : "))
        lst.append(input("Enter Phone : "))
        lst.append(input("Enter Age : "))
        lst.append(input("Enter Password : "))
        DataWrite = openpyxl.load_workbook(path)
        my_sheet = DataWrite["userData"]
        max_row = my_sheet.max_row
        for i in range(0, len(lst)):
            c = my_sheet.cell(row=max_row + 1, column=i + 1)
            c.value = lst[i]
        DataWrite.save(path)

    def Login(self, user, password):
        path = "data.xlsx"
        DataWrite = openpyxl.load_workbook(path)
        my_sheet = DataWrite["userData"]
        max_row = my_sheet.max_row
        for i in range(2, max_row + 1):
            data_user = my_sheet.cell(i, 1).value
            data_pass = my_sheet.cell(i, 5).value
            if ((data_user == user) and (data_pass == password)):
                return True

    def LoginFunctionalites(self):
        print("***Select a Movie***")
        with open('hello.csv') as f:
            firstColumn = [line.split(',')[0] for line in f]
        count=1
        for i in firstColumn:
            if i!="Title":
                print(count, i)
                count += 1
        choise=int(input("Enter the movie you want to choose : "))
        lst=[]
        lst1=[]
        with open('hello.csv', encoding="utf8") as f:
            csv_reader = csv.reader(f)
            for line in csv_reader:
                lst.append(line)
        option = lst[choise]
        print("Movie details :")
        print("Title - {} \nGenre - {} \nLength - {} \nCast - {} \nDirector - {} \nAdmin Rating - {} \nLanguage - {}\n".format(option[0],option[1],option[2],option[3],option[4],option[5],option[6]))
        numberofshoes=option[7]
        print("Select a show:")
        a=userOperations()
        with open('showTime.csv', encoding="utf8") as f:
            csv_reader = csv.reader(f)
            for line in csv_reader:
                lst1.append(line)
        cls1 = 1
        cls2 = 2
        select = lst1[choise]
        #print(select)
        while True:
            for i in range(int(numberofshoes)):
                print("{} Show time = {} ::: Tickets available {}".format(i + 1, select[cls1], select[cls2]))
                cls1 += 2
                cls2 += 2
            print("Which show you want to choose :")
            show = int(input())
            if show > int(numberofshoes):
                print("***Enter correct show***")
            else:
                print("Enter the operation you want to perform :\n1.Booking ticket \n2.cancel ticket\n3.logout")
                operation = int(input())
                if operation <= 3 and operation > 0:
                    match operation:
                        case 1:
                            print("Number of tickets you want to book :")
                            BookingTickets = int(input())
                            AvailableTicktet = select[show * 2]
                            if BookingTickets > int(AvailableTicktet):
                                print("Only {} tickets are available".format(AvailableTicktet))
                            else:
                                select[show * 2] = int(AvailableTicktet) - BookingTickets
                                a.editShow(select)
                                print(select)
                        case 2:
                            print("Number of ticktes you want to cancel :")
                            CancelTicktes = int(input())
                            AvailableTicktet = int(select[show * 2])
                            select[show * 2] = AvailableTicktet + CancelTicktes
                            a.editShow(select)
                        case 3:
                            print("Logged out successfully")
                            break
                else:
                    print("***Enter the correct option***")

    def editShow(self,arr):
        lines = list()
        with open('showTime.csv', 'r', newline='') as readFile:
            reader = csv.reader(readFile)
            for row in reader:
                lines.append(row)
                for field in row:
                    if field == arr[0]:
                        lst = row
                        lines.remove(row)
                        lst = arr
                        lines.append(lst)
        with open('showTime.csv', 'w', newline='') as writeFile:
            writer = csv.writer(writeFile)
            writer.writerows(lines)